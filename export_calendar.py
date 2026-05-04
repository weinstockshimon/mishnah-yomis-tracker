import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl


MONTH_DATE_RE = re.compile(r"\b(\w+)\s+(\d+)(?:st|nd|rd|th)\s+(\d{4})\b")


def parse_english_date(value):
    match = MONTH_DATE_RE.search(str(value or ""))
    if not match:
        raise ValueError(f"Could not parse English date: {value!r}")
    month, day, year = match.groups()
    parsed = datetime.strptime(f"{month} {day} {year}", "%B %d %Y")
    return parsed.date().isoformat()


def clean(value):
    return "" if value is None else str(value).strip()


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    source_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    workbook = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
    worksheet = workbook["print"] if "print" in workbook.sheetnames else workbook.active

    rows = []
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if not any(row[:5]):
            continue

        flag = clean(row[5] if len(row) > 5 else None)
        iso_date = parse_english_date(row[1])
        rows.append(
            {
                "id": iso_date,
                "row": row_number,
                "weekday": clean(row[0]),
                "englishDate": clean(row[1]),
                "isoDate": iso_date,
                "hebrewDate": clean(row[2]),
                "tractate": clean(row[3]),
                "assignment": clean(row[4]),
                "sourceFlag": flag,
                "sourceCompleted": bool(flag),
            }
        )

    payload = {
        "sourceWorkbook": source_path.name,
        "sourceSheet": worksheet.title,
        "flagColumn": "F",
        "flaggedRows": sum(1 for row in rows if row["sourceCompleted"]),
        "totalRows": len(rows),
        "items": rows,
    }

    output_path.write_text(
        "window.MISHNAH_YOMIS_DATA = "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )

    print(
        json.dumps(
            {
                "output": str(output_path),
                "rows": len(rows),
                "flagged": payload["flaggedRows"],
                "firstIncomplete": next(
                    (row for row in rows if not row["sourceCompleted"]), None
                ),
            },
            ensure_ascii=False,
            default=str,
        )
    )


if __name__ == "__main__":
    main()
